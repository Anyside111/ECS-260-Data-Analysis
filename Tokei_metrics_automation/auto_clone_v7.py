import os
import openpyxl
from git import Repo
from tqdm import tqdm
import calendar
from datetime import datetime, timedelta
import pandas as pd
import logging
import csv
import sys

def max_rows(sheet):
    count = 0
    for row in sheet.iter_rows(min_row=2):
        if(row[0].value):
            count= count + 1
    return count

def parse_hrlog_write_csv(file_name):
    # Split the log text into lines
    lines = [line.strip() for line in open(f"{file_name}.txt", "r").readlines()]

    # Identify the data header row
    header_row = None
    header = None
    for line in lines:
        if line.startswith("Language"):
            header_row = line.split("!")[0].strip().split()
            header = line
            break

    # Extract data from each line
    data = []
    parent=""
    for line in lines[lines.index(header) + 1 :]:
        if line:
            data_row = line.split("!")[0].strip().split() + line.split("!")[1:]

            if(data_row[0] == "|-"):
                data_row[1] = parent + "_" + data_row[1]
                data_row.remove("|-")
            elif(data_row[0] == "(Total)"):
                data_row[0] = parent + "_" + "Total"
                data_row.insert(1, "-")
            else:
                parent = data_row[0]

            if(len(data_row)>6):
                data_row[0] = data_row[0] + " " + data_row[1]
                data_row.remove(data_row[1])

            if(len(data_row)>1):
                data.append(data_row)

    # Write data to CSV file
    with open(f"{file_name}.csv", "w", newline="") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(header_row)
        writer.writerows(data)

    return f"{file_name}.csv"

def get_month_wise_ranges(start_date: datetime, end_date: datetime) -> list[tuple[datetime, datetime]]:
  """
  This function takes two datetime objects and returns a list of tuples,
  where each tuple represents a month-wise range with daily precision.

  Args:
      start_date: The starting datetime object (inclusive).
      end_date: The ending datetime object (inclusive).

  Returns:
      A list of tuples representing month-wise ranges with daily precision.
  """
  date_range = []
  current_date = start_date
  while current_date <= end_date:
    # Get last day of the current month
    last_day = current_date + timedelta(days=calendar.monthrange(current_date.year, current_date.month)[1] - 1)

    # Get the first day of the next month or the end date (whichever comes first)
    if(current_date.month == 12):
        next_month_start = current_date.replace(day=1, month=1, year = current_date.year + 1)
    else:
        next_month_start = current_date.replace(day=1, month=current_date.month + 1)
    next_date = min(next_month_start, end_date)

    # Append the current month range as a tuple
    date_range.append(current_date)

    # Update the current date to the first day of the next month
    current_date = next_month_start

  if end_date not in date_range:
    date_range.append(end_date)

  date_range.reverse()

  return date_range

def create_logger():
    # Configure a basic logger
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)  # Set the logging level (e.g., DEBUG, INFO, WARNING)

    # Define a handler to send logs to the console
    handler = logging.StreamHandler()
    handler.setLevel(logging.DEBUG)  # Set the handler's logging level

    # Create a formatter to format log messages
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)

    # Add the handler to the logger
    logger.addHandler(handler)

    # Use the logger to capture messages at different levels
    '''logger.debug('This is a debug message')
    logger.info('This is an informational message')
    logger.warning('This is a warning message')
    logger.error('This is an error message')
    logger.critical('This is a critical message')'''
    
    return logger


logger = create_logger()


# Define temporary folder for cloning
TEMP_DIR = f"temp_repos_{(datetime.now()).strftime('%m_%d_%H_%M')}"

FILENAME = "ASFI_list.xlsx"

# Create the temporary folder if it doesn't exist
os.makedirs(TEMP_DIR, exist_ok=True)

# Open the Excel file
wb = openpyxl.load_workbook(FILENAME)
sheet = wb.active

# Total number of repositories
total_repos = max_rows(sheet)

# Initial number of successfully cloned (starts at 0)
repo_count = 0

# Progress bar with dynamic title
progress_bar = tqdm(total=total_repos, desc="Cloning repositories (0/0)")

df_overall = pd.DataFrame()

df_granular = pd.DataFrame()

# Update progress bar title with initial successful count
progress_bar.set_description(f"Processing repositories ({repo_count}/{total_repos})")
progress_bar.update()

# Iterate through each row in the Excel sheet
for data_entry in sheet.iter_rows(min_row=2):
    # Get repository URL and end date
    repo_links = data_entry[7].value
    end_date_str = data_entry[3].value
    start_date_str = data_entry[2].value

    if(repo_links == None):
        continue

    df_row = {}
    df2_row = {}

    logger.info(f"Selected repo: {data_entry[1].value}")

    try:
        repo_list = repo_links.split("|")
        logger.info(f"Repo list for project: {repo_list}")
        repo_gits = []
        
        for repo_url in repo_list:
            
            if(repo_url == None):
                continue

            try:
                # Create repository name based on URL
                repo_name = repo_url.split("/")[-1]

                logger.info(f"Start time = {start_date_str}; end_date = {end_date_str}")

                end_date = end_date_str #datetime.fromisoformat(end_date_str)
                start_date = start_date_str #datetime.fromisoformat(start_date_str)

                print(f"\nRepo name = {repo_name}")
                #print(f"end_date: {end_date}")

                # Create the temporary folder if it doesn't exist
                os.makedirs(os.path.join(TEMP_DIR, data_entry[1].value), exist_ok=True)

                # Clone the repository with specific commit based on end date
                repo_dir = os.path.join(TEMP_DIR, data_entry[1].value, repo_name)
                repo = Repo.clone_from(repo_url, repo_dir, multi_options=['--config core.protectNTFS=false'], allow_unsafe_options=True)
                repo_gits.append(repo)
                #os.system(f"")
                #git = repo.git
                #git.config("config core.protectNTFS false")

            except Exception as e:
                logger.error(f"Exception received in first try block: {e}")
                continue

        date_range = get_month_wise_ranges(start_date, end_date)

        #logger.info(f"Date range calculated: {date_range}")

        successful_months_processed = 0

        for target_date in date_range:
            clone_ts_list =[]
            for repo in repo_gits:

                # Get the repository name from the working directory
                repo_name = os.path.basename(repo.working_dir)

                print(f"Repository Name: {repo_name}")

                commit = repo.head.commit

                #logger.info(f"target date: {target_date}; timezone info = {target_date.tzinfo}")
                #logger.info(f"commit date: {commit.committed_datetime}; timezone info = {commit.committed_datetime.tzinfo}")

                while commit.committed_datetime.replace(tzinfo=None) > target_date:
                    try:
                        commit = commit.parents[0]
                    except:
                        break
                
                #repo.git.checkout(commit.hexsha)
                try:
                    repo.head.reference = commit
                    #assert not repo.head.is_detached
                    repo.head.reset(index=True, working_tree=True)
                    
                    '''new_head = repo.create_head(f"new_head_{commit.hexsha}", commit)
                    repo.head.reference = new_head
                    repo.heads.new_head.checkout()'''

                    '''new_head = repo.create_head("new_head", commit)
                    repo.head.reference = new_head
                    repo.head.reset(index=True, working_tree=True)'''

                except Exception as e:
                    logger.info(f"error when resetting head: {e}")
                    df2_row = {"listid":data_entry[0].value, "listname": data_entry[1].value, "github_link": data_entry[7].value, "target_date": target_date, "Clone_status": "Failed", "clone_date": clone_date, "repo_dir": repo_dir, "metric_processing": "Failed", "log_dir": log_path}
                    df2_row = {k:[v] for k,v in df2_row.items()}
                    new_row = pd.DataFrame(df2_row)
                    df_granular = pd.concat([df_granular, new_row], ignore_index=True)
                    logger.info(f"should have appended: {new_row}")
                    continue

                commit = repo.head.commit

                clone_date = commit.committed_datetime.replace(tzinfo=None)

                logger.info(f"cloned timestamp = {clone_date}")
                clone_ts_list.append(f"{clone_date}")
                #logger.info(f"processing commit with hexsha: {commit.hexsha}")
                clone_date = ' | '.join(clone_ts_list)

            try:
                # Create the temporary folder if it doesn't exist
                log_path = os.path.join(TEMP_DIR, "_Autoclone_logs")
                os.makedirs(log_path, exist_ok=True)
                os.system(f"tokei ./{os.path.join(TEMP_DIR, data_entry[1].value)} > {os.path.join(log_path, data_entry[1].value)}_{target_date.year}_{target_date.month}_{target_date.day}_hrlog.txt")
                os.system(f"tokei -o json ./{os.path.join(TEMP_DIR, data_entry[1].value)} > {os.path.join(log_path, data_entry[1].value)}_{target_date.year}_{target_date.month}_{target_date.day}_jsonlog.txt")
                parse_hrlog_write_csv(f"{os.path.join(log_path, data_entry[1].value)}_{target_date.year}_{target_date.month}_{target_date.day}_hrlog")
        
                df2_row = {"listid":data_entry[0].value, "listname": data_entry[1].value, "github_link": data_entry[7].value, "target_date": target_date, "Clone_status": "Success", "clone_date": clone_date, "repo_dir": repo_dir, "metric_processing": "Success", "log_dir": log_path}
                successful_months_processed = successful_months_processed + 1
            except Exception as e:
                logger.info(f"Exception received in second try block: {e}")
                df2_row = {"listid":data_entry[0].value, "listname": data_entry[1].value, "github_link": data_entry[7].value, "target_date": target_date, "Clone_status": "Success", "clone_date": clone_date, "repo_dir": repo_dir, "metric_processing": "Failed", "log_dir": log_path}

            df2_row = {k:[v] for k,v in df2_row.items()}
            new_row = pd.DataFrame(df2_row)
            df_granular = pd.concat([df_granular, new_row], ignore_index=True)
        # Log successful cloning and update variables
        if(successful_months_processed == len(date_range)):
            df_row = {"listid":data_entry[0].value, "listname": data_entry[1].value, "github_link": data_entry[7].value, "start_date": data_entry[2].value, "end_date": data_entry[3].value, "Clone_status": "Success", "path": repo_dir, "metric_processing": "Success", "log_dir": log_path}
        else:
            df_row = {"listid":data_entry[0].value, "listname": data_entry[1].value, "github_link": data_entry[7].value, "start_date": data_entry[2].value, "end_date": data_entry[3].value, "Clone_status": "Success", "path": repo_dir, "metric_processing": f"Failed {successful_months_processed}/{len(date_range)}", "log_dir": log_path}
            
        repo_count += 1

        # Update progress bar title with current successful count
        progress_bar.set_description(f"Processing repositories ({repo_count}/{total_repos})")

    except Exception as e:
        logger.info(f"Exception received in overall try: {e}")
        # Get the traceback object
        exc_type, exc_obj, exc_tb = sys.exc_info()
        # Extract the line number from the traceback
        line_number = exc_tb.tb_lineno
        print(f"Error on line {line_number}: {exc_type} - {exc_obj}")

        # Log failed link and error message
        df_row = {"listid":data_entry[0].value, "listname": data_entry[1].value, "github_link": data_entry[7].value, "start_date": data_entry[2].value, "end_date": data_entry[3].value, "Clone_status": "Failed", "path": None, "metric_processing": "Failed", "log_dir": None}

    # new_row = pd.Series(row, index="index")
    # df= df.concat([df, new_row], ignore_index=False, index=[repo_count])
    df_row = {k:[v] for k,v in df_row.items()}
    new_row = pd.DataFrame(df_row)
    df_overall = pd.concat([df_overall, new_row], ignore_index=True)

    # Update progress bar even on exceptions
    progress_bar.update()

# Close progress bar and log files
progress_bar.close()
#success_log.close()
#failed_links.close()

html_df = df_overall.to_html(index=False)
with open("auto_clone_status_overall.html", "w") as f:
    f.write(html_df)

df_overall.to_csv("auto_clone_status_overall.csv", index=False)

html_df = df_granular.to_html(index=False)
with open("auto_clone_status_granular.html", "w") as f:
    f.write(html_df)

df_granular.to_csv("auto_clone_status_granular.csv", index=False)

print("Cloning completed. Check auto_clone_status.html or auto_clone_status.csv for details.")        